from fastapi import FastAPI, Request, HTTPException, UploadFile, File
from fastapi.responses import RedirectResponse, HTMLResponse, FileResponse, PlainTextResponse, JSONResponse
import os
import re
import time
import subprocess
import csv
from io import BytesIO
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import urlencode, quote
import secrets
from psycopg2.extras import Json, RealDictCursor, execute_values

import requests

from etl.inventory.repository import db_connect

app = FastAPI()
#.

ML_CLIENT_ID = os.environ["ML_CLIENT_ID"]
ML_CLIENT_SECRET = os.environ["ML_CLIENT_SECRET"]
ML_REDIRECT_URI = os.environ["ML_REDIRECT_URI"]
GOOGLE_CLIENT_ID = os.environ["GOOGLE_CLIENT_ID"]
GOOGLE_CLIENT_SECRET = os.environ["GOOGLE_CLIENT_SECRET"]
GOOGLE_REDIRECT_URI = os.environ["GOOGLE_REDIRECT_URI"]
APP_SESSION_COOKIE_NAME = os.environ.get("APP_SESSION_COOKIE_NAME", "exos_saas_session")
APP_SESSION_DAYS = int(os.environ.get("APP_SESSION_DAYS", "30"))

AUTH_URL = "https://auth.mercadolivre.com.br/authorization"
TOKEN_URL = "https://api.mercadolibre.com/oauth/token"
ME_URL = "https://api.mercadolibre.com/users/me"
GOOGLE_AUTH_URL = "https://accounts.google.com/o/oauth2/v2/auth"
GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"
GOOGLE_USERINFO_URL = "https://openidconnect.googleapis.com/v1/userinfo"

BASE_DIR = Path("/opt/render/project/src")
CSV_DIR = BASE_DIR / "runtime_csv"
LOG_DIR = BASE_DIR / "runtime_logs"
CSV_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(parents=True, exist_ok=True)

ACTIVE_JOB_STATUSES = ("queued", "running")


LIVE_SUBSCRIPTION_STATUSES = ("trialing", "active", "past_due", "paused")

try:
    import openpyxl
except Exception:
    openpyxl = None



def ensure_trial_for_account(account_id: int) -> bool:
    """
    Retorna True se criou trial agora.
    Retorna False se já existia assinatura válida.
    """
    with db_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT 1
                FROM billing.subscription
                WHERE account_id = %s
                  AND status IN ('trialing', 'active', 'past_due', 'paused')
                LIMIT 1
                """,
                (account_id,),
            )
            if cur.fetchone():
                conn.commit()
                return False

            cur.execute(
                """
                SELECT id, trial_days
                FROM billing.plan
                WHERE code = 'trial'
                  AND is_active = TRUE
                LIMIT 1
                """
            )
            plan = cur.fetchone()
            if not plan:
                conn.commit()
                return False

            plan_id, trial_days = plan

            cur.execute(
                """
                INSERT INTO billing.subscription (
                    account_id,
                    plan_id,
                    provider,
                    status,
                    current_period_start,
                    current_period_end,
                    cancel_at_period_end,
                    created_at,
                    updated_at
                )
                VALUES (
                    %s,
                    %s,
                    'internal',
                    'trialing',
                    now(),
                    now() + (%s || ' days')::interval,
                    FALSE,
                    now(),
                    now()
                )
                """,
                (account_id, plan_id, int(trial_days or 10)),
            )

        conn.commit()

    return True

def get_live_subscription(account_id: int) -> dict | None:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    s.id,
                    s.account_id,
                    s.plan_id,
                    s.provider,
                    s.provider_customer_id,
                    s.provider_subscription_id,
                    s.status,
                    s.current_period_start,
                    s.current_period_end,
                    s.cancel_at_period_end,
                    p.code AS plan_code,
                    p.name AS plan_name,
                    p.price_monthly,
                    p.mlb_limit,
                    p.daily_execution_limit,
                    p.trial_days
                FROM billing.subscription s
                JOIN billing.plan p
                  ON p.id = s.plan_id
                WHERE s.account_id = %s
                  AND s.status IN ('trialing', 'active', 'past_due', 'paused')
                ORDER BY s.created_at DESC
                LIMIT 1
                """,
                (account_id,),
            )
            row = cur.fetchone()
    return dict(row) if row else None


def expire_trial_if_needed(subscription: dict | None) -> dict | None:
    if not subscription or subscription.get("status") != "trialing":
        return subscription

    period_end = subscription.get("current_period_end")
    if not period_end:
        return subscription

    now = datetime.now(timezone.utc)
    if period_end <= now:
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE billing.subscription
                    SET status = 'expired',
                        updated_at = now()
                    WHERE id = %s
                    """,
                    (subscription["id"],),
                )
            conn.commit()
        subscription["status"] = "expired"
    return subscription


def create_trial_subscription_if_missing(account_id: int) -> dict | None:
    current = get_live_subscription(account_id)
    if current:
        return current

    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, code, name, trial_days
                FROM billing.plan
                WHERE code = 'trial'
                  AND is_active = TRUE
                LIMIT 1
                """
            )
            trial_plan = cur.fetchone()

            if not trial_plan:
                conn.commit()
                return None

            cur.execute(
                """
                INSERT INTO billing.subscription (
                    account_id,
                    plan_id,
                    provider,
                    status,
                    current_period_start,
                    current_period_end,
                    cancel_at_period_end,
                    created_at,
                    updated_at
                )
                VALUES (
                    %s,
                    %s,
                    'internal',
                    'trialing',
                    now(),
                    now() + (%s || ' days')::interval,
                    FALSE,
                    now(),
                    now()
                )
                RETURNING id
                """,
                (account_id, trial_plan["id"], int(trial_plan["trial_days"] or 10)),
            )
            cur.fetchone()
        conn.commit()

    return get_live_subscription(account_id)


def get_or_create_subscription(account_id: int) -> dict | None:
    subscription = get_live_subscription(account_id)
    if subscription:
        subscription = expire_trial_if_needed(subscription)
        if subscription and subscription.get("status") == "expired":
            return None
        return subscription

    created = create_trial_subscription_if_missing(account_id)
    if not created:
        return None

    created = expire_trial_if_needed(created)
    if created and created.get("status") == "expired":
        return None
    return created


def get_today_usage(account_id: int) -> dict:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT account_id, usage_date, executions_count, last_execution_at
                FROM billing.usage_daily
                WHERE account_id = %s
                  AND usage_date = CURRENT_DATE
                """,
                (account_id,),
            )
            row = cur.fetchone()
    return dict(row) if row else {
        "account_id": account_id,
        "usage_date": None,
        "executions_count": 0,
        "last_execution_at": None,
    }


def validate_plan_access(account_id: int, requested_limit: int, enforce_mlb_limit: bool = True) -> dict:
    """
    Bloqueia qualquer passo caso:
      - não exista assinatura ativa/trial
      - trial tenha expirado
      - limite diário de execuções já tenha sido atingido

    Quando enforce_mlb_limit=True:
      - requested_limit precisa respeitar o mlb_limit do plano
      - requested_limit == 0 é bloqueado em planos com limite definido

    Quando enforce_mlb_limit=False:
      - não trava pela quantidade de MLBs, útil para inventory
    """
    subscription = get_or_create_subscription(account_id)
    if not subscription:
        raise HTTPException(
            status_code=403,
            detail="Sua conta não possui assinatura ativa. Escolha um plano para continuar.",
        )

    subscription = expire_trial_if_needed(subscription)
    if subscription and subscription.get("status") == "expired":
        raise HTTPException(
            status_code=403,
            detail="Seu período de teste expirou. Escolha um plano para continuar.",
        )

    plan_code = subscription["plan_code"]
    mlb_limit = subscription["mlb_limit"]
    daily_limit = int(subscription["daily_execution_limit"] or 0)

    if enforce_mlb_limit and mlb_limit is not None:
        if requested_limit == 0:
            raise HTTPException(
                status_code=403,
                detail=f"Seu plano {plan_code} não permite executar em todos os anúncios. Limite máximo: {mlb_limit} MLBs por execução.",
            )
        if requested_limit > int(mlb_limit):
            raise HTTPException(
                status_code=403,
                detail=f"Seu plano {plan_code} permite no máximo {mlb_limit} MLBs por execução.",
            )

    today_usage = get_today_usage(account_id)
    executions_today = int(today_usage.get("executions_count") or 0)
    if executions_today >= daily_limit:
        raise HTTPException(
            status_code=403,
            detail=f"Seu plano {plan_code} permite {daily_limit} execução(ões) por dia. Limite de hoje já foi atingido.",
        )

    return {
        "subscription": subscription,
        "usage_today": today_usage,
        "executions_remaining_today": max(daily_limit - executions_today, 0),
    }

def register_daily_execution(account_id: int) -> dict:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                INSERT INTO billing.usage_daily (
                    account_id,
                    usage_date,
                    executions_count,
                    last_execution_at,
                    created_at,
                    updated_at
                )
                VALUES (%s, CURRENT_DATE, 1, now(), now(), now())
                ON CONFLICT (account_id, usage_date)
                DO UPDATE SET
                    executions_count = billing.usage_daily.executions_count + 1,
                    last_execution_at = now(),
                    updated_at = now()
                RETURNING account_id, usage_date, executions_count, last_execution_at
                """,
                (account_id,),
            )
            row = cur.fetchone()
        conn.commit()
    return dict(row)



def ensure_account_sku_min_receive_table() -> None:
    with db_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE SCHEMA IF NOT EXISTS app;
                CREATE TABLE IF NOT EXISTS app.account_sku_min_receive (
                    id BIGSERIAL PRIMARY KEY,
                    account_id BIGINT NOT NULL,
                    sku TEXT NOT NULL,
                    vlr_min_receber NUMERIC(18,2) NOT NULL,
                    source_file_name TEXT,
                    uploaded_by_user_id BIGINT,
                    created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                    UNIQUE (account_id, sku)
                );
                CREATE INDEX IF NOT EXISTS idx_account_sku_min_receive_account
                ON app.account_sku_min_receive (account_id);
                """
            )
        conn.commit()


def parse_decimal_br(value) -> float:
    text = str(value or '').strip()
    if not text:
        raise ValueError('Valor vazio')
    if ',' in text:
        text = text.replace('.', '').replace(',', '.')
    return float(text)


def parse_min_receive_file(filename: str, content: bytes) -> list[tuple[str, float]]:
    name = (filename or '').lower()
    rows: list[tuple[str, float]] = []

    if name.endswith('.csv'):
        decoded = content.decode('utf-8-sig')
        sample = decoded[:4096]
        delim = ';' if sample.count(';') >= sample.count(',') else ','
        reader = csv.DictReader(decoded.splitlines(), delimiter=delim)
        fields = {str(f).strip().lower(): f for f in (reader.fieldnames or [])}
        if 'sku' not in fields or 'vlr_min_receber' not in fields:
            raise HTTPException(status_code=400, detail='O arquivo deve conter as colunas sku e vlr_min_receber.')
        for row in reader:
            sku = str(row.get(fields['sku']) or '').strip()
            if not sku:
                continue
            rows.append((sku, parse_decimal_br(row.get(fields['vlr_min_receber']))))
        return rows

    if name.endswith('.xlsx'):
        if openpyxl is None:
            raise HTTPException(status_code=500, detail='Suporte a XLSX não disponível. Adicione openpyxl ao requirements.txt.')
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        header = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if 'sku' not in header or 'vlr_min_receber' not in header:
            raise HTTPException(status_code=400, detail='O arquivo deve conter as colunas sku e vlr_min_receber.')
        idx_sku = header.index('sku')
        idx_val = header.index('vlr_min_receber')
        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = str(row[idx_sku] or '').strip()
            if not sku:
                continue
            rows.append((sku, parse_decimal_br(row[idx_val])))
        return rows

    raise HTTPException(status_code=400, detail='Formato inválido. Envie um arquivo .csv ou .xlsx.')


def replace_account_min_receive(account_id: int, uploaded_by_user_id: int, filename: str, rows: list[tuple[str, float]]) -> int:
    ensure_account_sku_min_receive_table()
    with db_connect() as conn:
        with conn.cursor() as cur:
            cur.execute('DELETE FROM app.account_sku_min_receive WHERE account_id = %s', (account_id,))
            if rows:
                execute_values(
                    cur,
                    """
                    INSERT INTO app.account_sku_min_receive (
                        account_id, sku, vlr_min_receber, source_file_name, uploaded_by_user_id, created_at, updated_at
                    ) VALUES %s
                    """,
                    [(account_id, sku, value, filename, uploaded_by_user_id) for sku, value in rows],
                    template='(%s,%s,%s,%s,%s,now(),now())',
                    page_size=500,
                )
        conn.commit()
    return len(rows)


def list_account_min_receive(account_id: int, limit: int = 50) -> list[dict]:
    ensure_account_sku_min_receive_table()
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT sku, vlr_min_receber, source_file_name, updated_at
                FROM app.account_sku_min_receive
                WHERE account_id = %s
                ORDER BY updated_at DESC, sku
                LIMIT %s
                """,
                (account_id, limit),
            )
            rows = cur.fetchall()
    return [dict(r) for r in rows]


def count_account_min_receive(account_id: int) -> int:
    ensure_account_sku_min_receive_table()
    with db_connect() as conn:
        with conn.cursor() as cur:
            cur.execute('SELECT count(*) FROM app.account_sku_min_receive WHERE account_id = %s', (account_id,))
            return int(cur.fetchone()[0] or 0)


def get_session_user(request: Request) -> dict | None:
    token = request.cookies.get(APP_SESSION_COOKIE_NAME)
    if not token:
        return None

    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    u.id,
                    u.email,
                    u.full_name,
                    u.google_sub,
                    u.picture_url,
                    u.status
                FROM app.web_session s
                JOIN app.user_account u
                  ON u.id = s.user_account_id
                WHERE s.session_token = %s
                  AND s.expires_at > now()
                  AND u.status = 'active'
                """,
                (token,),
            )
            row = cur.fetchone()
    return dict(row) if row else None


def require_user(request: Request) -> dict:
    user = get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Você precisa fazer login para acessar esta área.")
    return user


def get_user_account_ids(user_account_id: int) -> list[int]:
    with db_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT account_id
                FROM app.account_user
                WHERE user_account_id = %s
                """,
                (user_account_id,),
            )
            rows = cur.fetchall()
    return [int(r[0]) for r in rows]


def get_user_role_for_account(user_account_id: int, account_id: int) -> str | None:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT role
                FROM app.account_user
                WHERE user_account_id = %s
                  AND account_id = %s
                LIMIT 1
                """,
                (user_account_id, account_id),
            )
            row = cur.fetchone()
    return row["role"] if row else None


def require_account_role(user_account_id: int, account_id: int, allowed_roles: tuple[str, ...] = ("owner", "admin")) -> str:
    role = get_user_role_for_account(user_account_id, account_id)
    if not role or role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Você não tem permissão para gerenciar acessos desta conta.")
    return role

def create_account_for_user(user_account_id: int, email: str, full_name: str | None = None) -> int:
    email = (email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="E-mail inválido para criar conta.")

    account_name = (full_name or "").strip()
    if not account_name:
        account_name = email.split("@")[0]

    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                INSERT INTO app.account (name, created_at, updated_at)
                VALUES (%s, now(), now())
                RETURNING id
                """,
                (account_name,),
            )
            account_id = int(cur.fetchone()["id"])

            cur.execute(
                """
                INSERT INTO app.account_user (account_id, user_account_id, role, created_at)
                VALUES (%s, %s, 'owner', now())
                ON CONFLICT (account_id, user_account_id) DO NOTHING
                """,
                (account_id, user_account_id),
            )

        conn.commit()

    return account_id


def list_account_invites(account_id: int) -> list[dict]:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, account_id, email, role, status, invited_at, accepted_at, created_at, updated_at
                FROM app.account_user_invite
                WHERE account_id = %s
                ORDER BY created_at DESC
                """,
                (account_id,),
            )
            rows = cur.fetchall()
    return [dict(r) for r in rows]


def create_or_update_account_invite(account_id: int, email: str, role: str) -> dict:
    email = (email or "").strip().lower()
    role = (role or "viewer").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Informe um e-mail válido.")
    if role not in ("owner", "admin", "viewer"):
        raise HTTPException(status_code=400, detail="Role inválido.")

    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                INSERT INTO app.account_user_invite (account_id, email, role, status, invited_at, created_at, updated_at)
                VALUES (%s, %s, %s, 'pending', now(), now(), now())
                ON CONFLICT (account_id, email) DO UPDATE
                SET role = EXCLUDED.role,
                    status = 'pending',
                    accepted_at = NULL,
                    updated_at = now()
                RETURNING id, account_id, email, role, status, invited_at, accepted_at, created_at, updated_at
                """,
                (account_id, email, role),
            )
            row = cur.fetchone()
        conn.commit()
    return dict(row)


def revoke_account_invite(account_id: int, invite_id: int) -> None:
    with db_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE app.account_user_invite
                SET status = 'revoked',
                    updated_at = now()
                WHERE id = %s
                  AND account_id = %s
                """,
                (invite_id, account_id),
            )
        conn.commit()


def get_accessible_connected_sellers(user_account_id: int) -> list[dict]:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    cs.id AS connected_seller_id,
                    cs.account_id,
                    cs.ml_user_id,
                    cs.seller_nickname,
                    cs.site_id,
                    cs.status,
                    cs.authorized_at
                FROM ml.connected_seller cs
                JOIN app.account_user au
                  ON au.account_id = cs.account_id
                WHERE au.user_account_id = %s
                ORDER BY cs.created_at DESC
                """,
                (user_account_id,),
            )
            rows = cur.fetchall()
    return [dict(r) for r in rows]


def get_default_connected_seller_id_for_user(user_account_id: int) -> int | None:
    sellers = get_accessible_connected_sellers(user_account_id)
    return int(sellers[0]["connected_seller_id"]) if sellers else None


def require_connected_seller_access(user_account_id: int, connected_seller_id: int) -> dict:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    cs.id AS connected_seller_id,
                    cs.account_id,
                    cs.ml_user_id,
                    cs.seller_nickname,
                    cs.site_id,
                    cs.status
                FROM ml.connected_seller cs
                JOIN app.account_user au
                  ON au.account_id = cs.account_id
                WHERE au.user_account_id = %s
                  AND cs.id = %s
                """,
                (user_account_id, connected_seller_id),
            )
            row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=403, detail="Você não tem acesso a esta conta.")
    return dict(row)


def require_run_access(user_account_id: int, run_id: str) -> dict:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT j.run_id, j.connected_seller_id, cs.account_id
                FROM app.async_job j
                JOIN ml.connected_seller cs
                  ON cs.id = j.connected_seller_id
                JOIN app.account_user au
                  ON au.account_id = cs.account_id
                WHERE au.user_account_id = %s
                  AND j.run_id = %s
                """,
                (user_account_id, run_id),
            )
            row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=403, detail="Você não tem acesso a este job.")
    return dict(row)


def upsert_user_account_from_google(profile: dict) -> dict:
    email = (profile.get("email") or "").strip().lower()
    full_name = profile.get("name")
    google_sub = profile.get("sub")
    picture_url = profile.get("picture")

    if not email or not google_sub:
        raise HTTPException(status_code=400, detail="Não foi possível identificar o usuário do Google.")

    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                INSERT INTO app.user_account (
                    email, full_name, google_sub, picture_url, status, created_at, updated_at
                )
                VALUES (%s, %s, %s, %s, 'active', now(), now())
                ON CONFLICT (email)
                DO UPDATE SET
                    full_name = EXCLUDED.full_name,
                    google_sub = EXCLUDED.google_sub,
                    picture_url = EXCLUDED.picture_url,
                    updated_at = now()
                RETURNING id, email, full_name, google_sub, picture_url, status
                """,
                (email, full_name, google_sub, picture_url),
            )
            row = cur.fetchone()
        conn.commit()

    return dict(row)



def auto_link_user_by_invite(user_account_id: int, email: str) -> list[int]:
    email = (email or "").strip().lower()
    if not email:
        return []

    linked_account_ids: list[int] = []

    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, account_id, role
                FROM app.account_user_invite
                WHERE lower(email) = lower(%s)
                  AND status = 'pending'
                ORDER BY id
                """,
                (email,),
            )
            invites = cur.fetchall()

            for invite in invites:
                account_id = int(invite["account_id"])
                role = invite["role"] or "owner"

                cur.execute(
                    """
                    INSERT INTO app.account_user (account_id, user_account_id, role, created_at)
                    VALUES (%s, %s, %s, now())
                    ON CONFLICT (account_id, user_account_id) DO NOTHING
                    """,
                    (account_id, user_account_id, role),
                )

                cur.execute(
                    """
                    UPDATE app.account_user_invite
                    SET status = 'accepted',
                        accepted_at = now(),
                        updated_at = now()
                    WHERE id = %s
                    """,
                    (invite["id"],),
                )

                linked_account_ids.append(account_id)

        conn.commit()

    return linked_account_ids

def create_web_session(user_account_id: int) -> str:
    token = secrets.token_urlsafe(32)
    expires_at = utc_now() + timedelta(days=APP_SESSION_DAYS)

    with db_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO app.web_session (session_token, user_account_id, expires_at, created_at)
                VALUES (%s, %s, %s, now())
                """,
                (token, user_account_id, expires_at),
            )
        conn.commit()
    return token


def delete_web_session(session_token: str | None) -> None:
    if not session_token:
        return
    with db_connect() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM app.web_session WHERE session_token = %s", (session_token,))
        conn.commit()


def render_login_page(error_message: str = "") -> str:
    error_html = f'<div class="login-error">{error_message}</div>' if error_message else ''
    return f"""
    <html>
    <head>
        <title>Entrar | Exos Tools</title>
        <meta name="viewport" content="width=device-width, initial-scale=1" />
        <style>
            * {{ box-sizing: border-box; }}
            body {{ margin: 0; min-height: 100vh; display:flex; align-items:center; justify-content:center; font-family: Arial, sans-serif; background: linear-gradient(180deg, #06122b 0%, #091a3f 100%); color: #fff; }}
            .card {{ width: min(92vw, 460px); background: rgba(255,255,255,0.08); border: 1px solid rgba(255,255,255,0.08); border-radius: 24px; padding: 28px; box-shadow: 0 10px 30px rgba(0,0,0,0.22); text-align:center; }}
            h1 {{ margin: 0 0 10px; font-size: 34px; }}
            p {{ color: #d8e3ff; line-height: 1.5; }}
            .btn {{ display:inline-block; margin-top: 18px; width:100%; border:none; border-radius:16px; padding:16px 18px; font-size:18px; font-weight:700; background:#fff; color:#111827; text-decoration:none; }}
            .login-error {{ margin-top: 14px; padding: 12px 14px; border-radius: 12px; background: rgba(239,68,68,.16); border:1px solid rgba(239,68,68,.28); color:#fecaca; }}
            .muted {{ margin-top:12px; font-size:13px; color:#9fb0d9; }}
        
</style>
    </head>
    <body>
        <div class="card">
            <h1>🚀 Exos Tools</h1>
            <p>Entre com sua conta Google para acessar sua área de otimização de campanhas do Mercado Livre.</p>
            <a class="btn" href="/auth/google/start">Entrar com Google</a>
            {error_html}
            <div class="muted">Acesso liberado somente para e-mails autorizados.</div>
        </div>
    </body>
    </html>
    """

@app.get("/onboarding", response_class=HTMLResponse)
def onboarding_page(request: Request):
    user = require_user(request)
    account_ids = get_user_account_ids(int(user["id"]))
    account_id = int(account_ids[0]) if account_ids else None
    connect_href = f"/ml/oauth/start?account_id={account_id}" if account_id else "/painel"

    return f"""
    <!DOCTYPE html>
    <html lang="pt-br">
    <head>
      <meta charset="UTF-8">
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
      <title>Bem-vindo ao EXOS Profit</title>
      <style>
        body {{
          margin: 0;
          font-family: Inter, Arial, sans-serif;
          background: #020617;
          color: #e2e8f0;
          min-height: 100vh;
          display: flex;
          align-items: center;
          justify-content: center;
          padding: 24px;
        }}
        .box {{
          width: 100%;
          max-width: 760px;
          text-align: center;
          background: rgba(255,255,255,.04);
          border: 1px solid rgba(255,255,255,.08);
          border-radius: 24px;
          padding: 40px 32px;
          box-shadow: 0 24px 60px rgba(0,0,0,.35);
        }}
        h1 {{
          margin: 0 0 14px;
          font-size: 42px;
          line-height: 1.05;
        }}
        .gradient {{
          background: linear-gradient(90deg,#2FD4C7,#38bdf8);
          -webkit-background-clip: text;
          -webkit-text-fill-color: transparent;
        }}
        p {{
          color: #94a3b8;
          font-size: 18px;
          line-height: 1.6;
          margin: 12px 0;
        }}
        .btn {{
          display: inline-block;
          margin-top: 24px;
          background: linear-gradient(135deg,#2FD4C7,#1CB3A6);
          color: #022c22;
          padding: 16px 28px;
          border-radius: 12px;
          text-decoration: none;
          font-weight: 800;
        }}
        .micro {{
          margin-top: 12px;
          font-size: 14px;
          color: #64748b;
        }}
      
</style>
    </head>
    <body>
      <div class="box">
        <h1>🎉 Seu teste grátis <span class="gradient">começou</span></h1>
        <p>Você tem <b>10 dias</b> para testar o EXOS Profit e descobrir oportunidades reais de lucro no Mercado Livre.</p>
        <p>Próximo passo: <b>conectar sua conta do Mercado Livre</b> para começar a análise.</p>
        <a class="btn" href="{connect_href}">Conectar conta do Mercado Livre</a>
        <div class="micro">Logado como: {user.get('email')}</div>
      </div>
    </body>
    </html>
    """

@app.get("/login", response_class=HTMLResponse)
def login_page(error: str | None = None):
    return render_login_page(error or "")


@app.get("/auth/google/start")
def auth_google_start(request: Request):
    state = secrets.token_urlsafe(24)
    return_to = request.query_params.get("return_to", "/painel")
    params = {
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": GOOGLE_REDIRECT_URI,
        "response_type": "code",
        "scope": "openid email profile",
        "access_type": "online",
        "include_granted_scopes": "true",
        "prompt": "select_account",
        "state": state,
    }
    response = RedirectResponse(f"{GOOGLE_AUTH_URL}?{urlencode(params)}")
    response.set_cookie("google_oauth_state", state, httponly=True, secure=True, samesite="lax", max_age=600)
    response.set_cookie("post_login_redirect", return_to, httponly=True, secure=True, samesite="lax", max_age=600)
    return response




@app.post("/auth/logout")
def auth_logout(request: Request):
    token = request.cookies.get(APP_SESSION_COOKIE_NAME)
    delete_web_session(token)
    response = JSONResponse({"ok": True})
    response.delete_cookie(APP_SESSION_COOKIE_NAME)
    return response


@app.get("/account/invites")
def account_invites(request: Request, account_id: int):
    user = require_user(request)
    require_account_role(int(user["id"]), account_id)
    invites = list_account_invites(account_id)
    return {"invites": invites}


@app.post("/account/invites/create")
def account_invites_create(request: Request, account_id: int, email: str, role: str = "viewer"):
    user = require_user(request)
    require_account_role(int(user["id"]), account_id)
    invite = create_or_update_account_invite(account_id, email, role)
    return {"ok": True, "invite": invite}


@app.post("/account/invites/revoke")
def account_invites_revoke(request: Request, account_id: int, invite_id: int):
    user = require_user(request)
    require_account_role(int(user["id"]), account_id)
    revoke_account_invite(account_id, invite_id)
    return {"ok": True}


def get_job_csv_payload(run_id: str) -> dict | None:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT run_id, csv_file, csv_content, csv_mime_type, csv_bytes
                FROM app.async_job
                WHERE run_id = %s
                """,
                (run_id,),
            )
            row = cur.fetchone()
    return dict(row) if row else None


def purge_old_finished_jobs(connected_seller_id: int, keep_run_id: str, keep_last: int = 1) -> int:
    """
    Mantém apenas os últimos jobs finalizados/erro do seller.
    Não mexe em queued/running.
    """
    with db_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                WITH ranked AS (
                    SELECT run_id,
                           row_number() OVER (
                               PARTITION BY connected_seller_id
                               ORDER BY created_at DESC
                           ) AS rn
                    FROM app.async_job
                    WHERE connected_seller_id = %s
                      AND status IN ('finished', 'error')
                      AND run_id <> %s
                )
                DELETE FROM app.async_job j
                USING ranked r
                WHERE j.run_id = r.run_id
                  AND r.rn > %s
                """,
                (connected_seller_id, keep_run_id, keep_last),
            )
            deleted = cur.rowcount
        conn.commit()
    return deleted

def build_csv_path(prefix: str = "campaign_optimizer_audit") -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return CSV_DIR / f"{prefix}_{timestamp}.csv"


def build_log_path(prefix: str = "pipeline_run") -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return LOG_DIR / f"{prefix}_{timestamp}.log"


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def fmt_dt(value) -> str:
    if not value:
        return "-"
    if isinstance(value, str):
        return value
    return value.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def get_connected_seller_summary(connected_seller_id: int) -> dict:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    cs.id,
                    cs.account_id,
                    cs.ml_user_id,
                    cs.seller_nickname,
                    cs.site_id,
                    cs.status,
                    cs.authorized_at,
                    ot.expires_at,
                    ot.updated_at AS token_updated_at
                FROM ml.connected_seller cs
                LEFT JOIN ml.oauth_token ot
                  ON ot.connected_seller_id = cs.id
                WHERE cs.id = %s
                """,
                (connected_seller_id,),
            )
            row = cur.fetchone()

    if not row:
        return {
            "exists": False,
            "connected": False,
            "connected_seller_id": connected_seller_id,
        }

    connected = bool(row["ml_user_id"]) and row["status"] == "active"

    return {
        "exists": True,
        "connected": connected,
        "connected_seller_id": row["id"],
        "account_id": row["account_id"],
        "ml_user_id": row["ml_user_id"],
        "seller_nickname": row["seller_nickname"],
        "site_id": row["site_id"],
        "status": row["status"],
        "authorized_at": row["authorized_at"],
        "expires_at": row["expires_at"],
        "token_updated_at": row["token_updated_at"],
    }


def get_active_job_for_seller(connected_seller_id: int) -> dict | None:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    run_id, job_type, status, step, connected_seller_id,
                    limit_items, dry_run, use_cost, log_file, csv_file,
                    (csv_content IS NOT NULL) AS has_csv,
                    error, created_at, started_at, finished_at, updated_at,
                    payload_json, result_json
                FROM app.async_job
                WHERE connected_seller_id = %s
                  AND status IN ('queued', 'running')
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (connected_seller_id,),
            )
            row = cur.fetchone()
    return dict(row) if row else None


def insert_job(
    *,
    job_type: str,
    connected_seller_id: int,
    limit_items: int,
    dry_run: bool | None,
    use_cost: bool | None,
    payload: dict,
    log_file: str | None,
    csv_file: str | None,
) -> str:
    active_job = get_active_job_for_seller(connected_seller_id)
    if active_job:
        raise HTTPException(
            status_code=409,
            detail={
                "message": f"Já existe um job em andamento para o seller {connected_seller_id}.",
                "run_id": active_job["run_id"],
                "job_type": active_job["job_type"],
                "status": active_job["status"],
                "step": active_job["step"],
            },
        )

    run_id = os.urandom(16).hex()
    with db_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO app.async_job (
                    run_id,
                    job_type,
                    status,
                    step,
                    connected_seller_id,
                    limit_items,
                    dry_run,
                    use_cost,
                    payload_json,
                    log_file,
                    csv_file,
                    result_json,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, 'queued', NULL, %s, %s, %s, %s, %s, %s, %s, %s, now(), now())
                """,
                (
                    run_id,
                    job_type,
                    connected_seller_id,
                    limit_items,
                    dry_run,
                    use_cost,
                    Json(payload),
                    log_file,
                    csv_file,
                    Json({}),
                ),
            )
        conn.commit()
    return run_id


def get_job(run_id: str) -> dict:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    run_id,
                    job_type,
                    status,
                    step,
                    connected_seller_id,
                    limit_items,
                    dry_run,
                    use_cost,
                    payload_json,
                    result_json,
                    log_file,
                    csv_file,
                    (csv_content IS NOT NULL) AS has_csv,
                    error,
                    created_at,
                    started_at,
                    finished_at,
                    updated_at
                FROM app.async_job
                WHERE run_id = %s
                """,
                (run_id,),
            )
            row = cur.fetchone()

    if not row:
        raise HTTPException(status_code=404, detail=f"run_id {run_id} não encontrado")

    out = dict(row)
    if out.get("payload_json") is None:
        out["payload_json"] = {}
    if out.get("result_json") is None:
        out["result_json"] = {}
    out["summary"] = build_job_summary(out)
    return out


def build_inventory_cmd(connected_seller_id: int, limit_items: int) -> list[str]:
    cmd = [
        "python3",
        "-m",
        "etl.ml_inventory_snapshot_basic",
        "--connected-seller-id",
        str(connected_seller_id),
    ]
    if limit_items and limit_items > 0:
        cmd.extend(["--limit-items", str(limit_items)])
    return cmd


def build_rebate_cmd(connected_seller_id: int, limit_items: int) -> list[str]:
    cmd = [
        "python3",
        "-m",
        "etl.ml_item_promo_rebate_snapshot",
        "--connected-seller-id",
        str(connected_seller_id),
    ]
    if limit_items and limit_items > 0:
        cmd.extend(["--limit-items", str(limit_items)])
    return cmd


def build_optimizer_cmd(
    connected_seller_id: int,
    limit_items: int,
    dry_run: bool,
    use_cost: bool,
    csv_path: Path,
) -> list[str]:
    cmd = [
        "python3",
        "-m",
        "etl.ml_campaign_optimizer",
        "--connected-seller-id",
        str(connected_seller_id),
        "--use-cost",
        "true" if use_cost else "false",
        "--out",
        str(csv_path),
    ]
    if limit_items and limit_items > 0:
        cmd.extend(["--limit", str(limit_items)])
    if dry_run:
        cmd.append("--dry-run")
    return cmd


def parse_optimizer_stats(text: str) -> dict:
    stats = {}
    if not text:
        return stats
    patterns = {
        "scope_items": r'"scope_items":\s*(\d+)|scope_items=(\d+)',
        "processed": r'"processed":\s*(\d+)|processed=(\d+)',
        "switched": r'"switched":\s*(\d+)|switched=(\d+)',
        "errors": r'"errors":\s*(\d+)|errors=(\d+)',
        "no_action": r'"no_action":\s*(\d+)|no_action=(\d+)',
    }
    for key, pattern in patterns.items():
        m = re.search(pattern, text)
        if m:
            val = next((g for g in m.groups() if g is not None), None)
            if val is not None:
                stats[key] = int(val)
    return stats


def parse_rebate_stats(text: str) -> dict:
    stats = {}
    if not text:
        return stats
    patterns = {
        "mlb_count": r"'mlb_count':\s*(\d+)|mlb_count=(\d+)",
        "scope_rows": r"'scope_rows':\s*(\d+)|scope_rows=(\d+)",
        "rows_inserted": r"'rows_inserted':\s*(\d+)|rows_inserted=(\d+)",
        "failed_items": r"'failed_items':\s*(\d+)|failed_items=(\d+)",
    }
    for key, pattern in patterns.items():
        m = re.search(pattern, text)
        if m:
            val = next((g for g in m.groups() if g is not None), None)
            if val is not None:
                stats[key] = int(val)
    return stats


def parse_inventory_stats(text: str) -> dict:
    stats = {}
    if not text:
        return stats
    patterns = {
        "item_ids_found": r"'item_ids_found':\s*(\d+)|IDs encontrados:\s*(\d+)",
        "rows_inserted": r"'rows_inserted':\s*(\d+)|rows_inserted=(\d+)",
    }
    for key, pattern in patterns.items():
        m = re.search(pattern, text)
        if m:
            val = next((g for g in m.groups() if g is not None), None)
            if val is not None:
                stats[key] = int(val)
    return stats


def build_job_summary(job: dict) -> dict:
    summary = {
        "headline": "Job enfileirado.",
        "status_label": job.get("status", "queued"),
        "details": [],
        "metrics": {},
    }

    status = job.get("status")
    job_type = job.get("job_type")
    step = job.get("step")
    result_json = job.get("result_json") or {}

    if status == "queued":
        summary["headline"] = "Job enfileirado, aguardando processamento."
        return summary

    if status == "running":
        summary["headline"] = f"Executando job de {job_type}."
        if step:
            summary["details"].append(f"Etapa atual: {step}")
        return summary

    if status == "error":
        summary["headline"] = "Job finalizado com erro."
        if job.get("error"):
            summary["details"].append(str(job["error"]))
        return summary

    if status == "finished":
        if job_type == "inventory":
            inv = parse_inventory_stats(((result_json.get("inventory") or {}).get("stdout") or ""))
            summary["headline"] = f"Inventory finalizado. {inv.get('item_ids_found', 0)} anúncios lidos."
            summary["metrics"] = inv
            return summary

        if job_type == "rebate":
            reb = parse_rebate_stats(((result_json.get("rebate") or {}).get("stdout") or ""))
            failed = reb.get("failed_items", 0)
            summary["headline"] = (
                f"Rebate finalizado. {reb.get('rows_inserted', 0)} linhas gravadas."
                if failed == 0 else
                f"Rebate finalizado com {failed} falhas pontuais."
            )
            summary["metrics"] = reb
            return summary

        if job_type == "optimizer":
            opt = parse_optimizer_stats(((result_json.get("optimizer") or {}).get("stdout") or ""))
            switched = opt.get("switched", 0)
            processed = opt.get("processed", 0)
            if switched > 0:
                summary["headline"] = f"Optimizer finalizado. {switched} campanhas elegíveis para troca em {processed} itens analisados."
            else:
                summary["headline"] = f"Optimizer finalizado. Nenhuma troca necessária em {processed} itens analisados."
            summary["metrics"] = opt
            return summary

        if job_type == "full":
            inv = parse_inventory_stats(((result_json.get("inventory") or {}).get("stdout") or ""))
            reb = parse_rebate_stats(((result_json.get("rebate") or {}).get("stdout") or ""))
            opt = parse_optimizer_stats(((result_json.get("optimizer") or {}).get("stdout") or ""))
            processed = opt.get("processed", 0)
            switched = opt.get("switched", 0)
            errors = opt.get("errors", 0)
            if switched > 0:
                summary["headline"] = f"Pipeline finalizada. {switched} campanhas elegíveis para troca em {processed} itens analisados."
            else:
                summary["headline"] = f"Pipeline finalizada. Nenhuma troca necessária em {processed} itens analisados."
            summary["metrics"] = {
                "item_ids_found": inv.get("item_ids_found"),
                "rebate_failed_items": reb.get("failed_items"),
                "processed": processed,
                "switched": switched,
                "errors": errors,
                "no_action": opt.get("no_action"),
            }
            if reb.get("failed_items", 0) > 0:
                summary["details"].append(f"Rebate terminou com {reb.get('failed_items')} falhas pontuais da API.")
            return summary

    return summary


def recent_jobs_rows(limit: int = 20, connected_seller_id: int | None = None) -> list[dict]:
    sql = """
    SELECT
        run_id, job_type, status, step, connected_seller_id,
        limit_items, dry_run, use_cost, payload_json, result_json,
        log_file, csv_file, (csv_content IS NOT NULL) AS has_csv, error, created_at, started_at, finished_at, updated_at
    FROM app.async_job
    """
    params = []
    if connected_seller_id is not None:
        sql += " WHERE connected_seller_id = %s"
        params.append(connected_seller_id)
    sql += " ORDER BY created_at DESC LIMIT %s"
    params.append(limit)

    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

    out = []
    for row in rows:
        d = dict(row)
        d["summary"] = build_job_summary(d)
        out.append(d)
    return out


def get_latest_inventory_item_count(connected_seller_id: int) -> int | None:
    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT job_type, result_json
                FROM app.async_job
                WHERE connected_seller_id = %s
                  AND status = 'finished'
                  AND job_type IN ('inventory', 'full')
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (connected_seller_id,),
            )
            row = cur.fetchone()

    if not row:
        return None

    result_json = row.get("result_json") or {}
    if row.get("job_type") == "inventory":
        stdout = ((result_json.get("inventory") or {}).get("stdout") or "")
    else:
        stdout = ((result_json.get("inventory") or {}).get("stdout") or "")

    stats = parse_inventory_stats(stdout)
    value = stats.get("item_ids_found")
    return int(value) if value is not None else None


def badge(status: str | None) -> str:
    status = (status or "").lower()

    color_map = {
        "queued": ("#1d4ed8", "#dbeafe"),
        "running": ("#a16207", "#fef3c7"),
        "finished": ("#166534", "#dcfce7"),
        "error": ("#991b1b", "#fee2e2"),
    }

    bg, fg = color_map.get(status, ("#334155", "#e2e8f0"))
    label = status.upper() if status else "N/A"

    return (
        f'<span style="display:inline-block; padding:4px 8px; '
        f'border-radius:999px; font-size:12px; font-weight:700; '
        f'background:{bg}; color:{fg};">{label}</span>'
    )

@app.get("/")
def root(request: Request):
    user = get_session_user(request)
    return RedirectResponse(url="/painel" if user else "/login")


@app.get("/template/sku-min-receber.csv")
def download_template_min_receive_csv(request: Request):
    require_user(request)
    headers = {"Content-Disposition": 'attachment; filename="template_sku_min_receber.csv"'}
    return PlainTextResponse("sku;vlr_min_receber\nSKU-EXEMPLO-1;120,00\nSKU-EXEMPLO-2;95,50\n", media_type="text/csv", headers=headers)


@app.get("/account/min-receive")
def account_min_receive_list(request: Request, account_id: int):
    user = require_user(request)
    require_account_role(int(user["id"]), account_id, ("owner", "admin", "viewer"))
    return {"count": count_account_min_receive(account_id), "rows": list_account_min_receive(account_id, 50)}


@app.post("/account/min-receive/upload")
async def account_min_receive_upload(request: Request, account_id: int, file: UploadFile = File(...)):
    user = require_user(request)
    require_account_role(int(user["id"]), account_id, ("owner", "admin"))
    content = await file.read()
    rows = parse_min_receive_file(file.filename or '', content)
    if not rows:
        raise HTTPException(status_code=400, detail='Nenhum registro válido encontrado no arquivo.')
    saved = replace_account_min_receive(account_id, int(user['id']), file.filename or 'upload', rows)
    return {"ok": True, "rows_saved": saved}


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/ml/oauth/start")
def start_oauth(
    connected_seller_id: int | None = None,
    account_id: int | None = None,
):
    if connected_seller_id:
        state = f"seller:{connected_seller_id}"
    else:
        if not account_id:
            raise HTTPException(status_code=400, detail="account_id é obrigatório para conectar uma nova conta.")
        state = f"seller:new:{account_id}"

    params = {
        "response_type": "code",
        "client_id": ML_CLIENT_ID,
        "redirect_uri": ML_REDIRECT_URI,
        "state": state,
    }
    return RedirectResponse(f"{AUTH_URL}?{urlencode(params)}")



@app.get("/ml/oauth/callback")
async def oauth_callback(request: Request):
    current_user = require_user(request)
    current_user_id = int(current_user["id"])

    code = request.query_params.get("code")
    state = request.query_params.get("state")

    if not code:
        raise HTTPException(status_code=400, detail="Missing code")
    if not state or not state.startswith("seller:"):
        raise HTTPException(status_code=400, detail="Invalid state")

    connected_seller_id: int | None = None
    account_id: int | None = None
    is_new = False

    parts = state.split(":")
    if len(parts) == 3 and parts[1] == "new":
        is_new = True
        account_id = int(parts[2])
    elif len(parts) == 2:
        connected_seller_id = int(parts[1])
    else:
        raise HTTPException(status_code=400, detail="Invalid state format")

    token_resp = requests.post(
        TOKEN_URL,
        data={
            "grant_type": "authorization_code",
            "client_id": ML_CLIENT_ID,
            "client_secret": ML_CLIENT_SECRET,
            "code": code,
            "redirect_uri": ML_REDIRECT_URI,
        },
        timeout=30,
    )
    if token_resp.status_code != 200:
        raise HTTPException(status_code=500, detail=f"Token error: {token_resp.text}")

    tokens = token_resp.json()
    access_token = tokens["access_token"]
    refresh_token = tokens["refresh_token"]
    token_type = tokens.get("token_type", "Bearer")
    scope = tokens.get("scope")
    expires_at = utc_now() + timedelta(seconds=int(tokens["expires_in"]))

    me_resp = requests.get(
        ME_URL,
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=30,
    )
    if me_resp.status_code != 200:
        raise HTTPException(status_code=500, detail=f"Failed to fetch user: {me_resp.text}")

    me = me_resp.json()
    ml_user_id = me["id"]
    seller_nickname = me.get("nickname")
    site_id = me.get("site_id")

    with db_connect() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            if is_new:
                cur.execute(
                    "SELECT id, account_id FROM ml.connected_seller WHERE ml_user_id = %s",
                    (ml_user_id,),
                )
                existing = cur.fetchone()

                if existing:
                    connected_seller_id = int(existing["id"])
                    account_id = int(existing["account_id"])

                    cur.execute(
                        """
                        INSERT INTO app.account_user (account_id, user_account_id, role, created_at)
                        VALUES (%s, %s, 'owner', now())
                        ON CONFLICT (account_id, user_account_id) DO NOTHING
                        """,
                        (account_id, current_user_id),
                    )

                    cur.execute(
                        """
                        UPDATE ml.connected_seller
                           SET seller_nickname = %s,
                               site_id = %s,
                               status = 'active',
                               authorized_at = now(),
                               updated_at = now()
                         WHERE id = %s
                        """,
                        (seller_nickname, site_id, connected_seller_id),
                    )
                else:
                    cur.execute(
                        """
                        INSERT INTO ml.connected_seller (
                            account_id, ml_user_id, seller_nickname, site_id, status, authorized_at, created_at, updated_at
                        )
                        VALUES (%s, %s, %s, %s, 'active', now(), now(), now())
                        RETURNING id
                        """,
                        (account_id, ml_user_id, seller_nickname, site_id),
                    )
                    connected_seller_id = int(cur.fetchone()["id"])

                    cur.execute(
                        """
                        INSERT INTO app.account_user (account_id, user_account_id, role, created_at)
                        VALUES (%s, %s, 'owner', now())
                        ON CONFLICT (account_id, user_account_id) DO NOTHING
                        """,
                        (account_id, current_user_id),
                    )
            else:
                cur.execute(
                    """
                    UPDATE ml.connected_seller
                       SET ml_user_id = %s,
                           seller_nickname = %s,
                           site_id = %s,
                           status = 'active',
                           authorized_at = now(),
                           updated_at = now()
                     WHERE id = %s
                    RETURNING id, account_id
                    """,
                    (ml_user_id, seller_nickname, site_id, connected_seller_id),
                )
                row = cur.fetchone()
                if not row:
                    raise HTTPException(status_code=404, detail=f"connected_seller_id {connected_seller_id} não encontrado")
                account_id = int(row["account_id"])

                cur.execute(
                    """
                    INSERT INTO app.account_user (account_id, user_account_id, role, created_at)
                    VALUES (%s, %s, 'owner', now())
                    ON CONFLICT (account_id, user_account_id) DO NOTHING
                    """,
                    (account_id, current_user_id),
                )

            cur.execute(
                """
                INSERT INTO ml.oauth_token (
                    connected_seller_id, access_token, refresh_token, token_type, scope, expires_at, last_refresh_at, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, now(), now())
                ON CONFLICT (connected_seller_id)
                DO UPDATE SET
                    access_token = EXCLUDED.access_token,
                    refresh_token = EXCLUDED.refresh_token,
                    token_type = EXCLUDED.token_type,
                    scope = EXCLUDED.scope,
                    expires_at = EXCLUDED.expires_at,
                    last_refresh_at = now(),
                    updated_at = now()
                """,
                (connected_seller_id, access_token, refresh_token, token_type, scope, expires_at),
            )
        conn.commit()

    redirect_url = f"/painel?connected_seller_id={connected_seller_id}&connected=1"
    if account_id:
        redirect_url += f"&account_id={account_id}"
    return RedirectResponse(url=redirect_url)

@app.get("/auth/google/callback")
def auth_google_callback(request: Request):
    code = request.query_params.get("code")
    state = request.query_params.get("state")
    state_cookie = request.cookies.get("google_oauth_state")
    return_to = request.cookies.get("post_login_redirect") or "/painel"

    if not code or not state or not state_cookie or state != state_cookie:
        return RedirectResponse(url="/login?error=" + quote("Falha na autenticação com Google."))

    token_resp = requests.post(
        GOOGLE_TOKEN_URL,
        data={
            "client_id": GOOGLE_CLIENT_ID,
            "client_secret": GOOGLE_CLIENT_SECRET,
            "code": code,
            "grant_type": "authorization_code",
            "redirect_uri": GOOGLE_REDIRECT_URI,
        },
        timeout=30,
    )

    if token_resp.status_code != 200:
        return RedirectResponse(url="/login?error=" + quote("Não foi possível concluir o login com Google."))

    access_token = token_resp.json().get("access_token")

    profile_resp = requests.get(
        GOOGLE_USERINFO_URL,
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=30,
    )

    user = upsert_user_account_from_google(profile_resp.json())

    auto_link_user_by_invite(int(user["id"]), user["email"])

    account_ids = get_user_account_ids(int(user["id"]))

    # 🔥 CORREÇÃO PRINCIPAL
    if not account_ids:
        account_id = create_account_for_user(
            user_account_id=int(user["id"]),
            email=user["email"],
            full_name=user.get("full_name"),
        )
    else:
        account_id = int(account_ids[0])

    trial_created_now = ensure_trial_for_account(account_id)

    session_token = create_web_session(int(user["id"]))

    redirect_url = "/onboarding" if trial_created_now else return_to

    response = RedirectResponse(url=redirect_url)
    response.set_cookie(
        APP_SESSION_COOKIE_NAME,
        session_token,
        httponly=True,
        secure=True,
        samesite="lax",
        max_age=APP_SESSION_DAYS * 24 * 60 * 60,
    )

    response.delete_cookie("google_oauth_state")
    response.delete_cookie("post_login_redirect")

    return response

@app.get("/run/inventory")
def run_inventory(request: Request, connected_seller_id: int = 1, limit: int = 0):
    user = require_user(request)
    require_connected_seller_access(int(user["id"]), connected_seller_id)
    started = time.time()
    cmd = build_inventory_cmd(connected_seller_id, limit)
    result = subprocess.run(cmd, capture_output=True, text=True)
    return {
        "status": "inventory executado",
        "connected_seller_id": connected_seller_id,
        "limit": limit,
        "elapsed_seconds": round(time.time() - started, 2),
        "cmd": cmd,
        "returncode": result.returncode,
        "stdout": result.stdout,
        "stderr": result.stderr,
        "csv_file": None,
    }


@app.get("/run/rebate")
def run_rebate(request: Request, connected_seller_id: int = 1, limit: int = 0):
    user = require_user(request)
    require_connected_seller_access(int(user["id"]), connected_seller_id)
    started = time.time()
    cmd = build_rebate_cmd(connected_seller_id, limit)
    result = subprocess.run(cmd, capture_output=True, text=True)
    return {
        "status": "rebate executado",
        "connected_seller_id": connected_seller_id,
        "limit": limit,
        "elapsed_seconds": round(time.time() - started, 2),
        "cmd": cmd,
        "returncode": result.returncode,
        "stdout": result.stdout,
        "stderr": result.stderr,
        "csv_file": None,
    }


@app.get("/run/optimizer")
def run_optimizer(request: Request, connected_seller_id: int = 1, limit: int = 0, dry_run: bool = True, use_cost: bool = False):
    user = require_user(request)
    require_connected_seller_access(int(user["id"]), connected_seller_id)
    started = time.time()
    csv_path = build_csv_path()
    cmd = build_optimizer_cmd(connected_seller_id, limit, dry_run, use_cost, csv_path)
    result = subprocess.run(cmd, capture_output=True, text=True)
    return {
        "status": "optimizer executado",
        "connected_seller_id": connected_seller_id,
        "limit": limit,
        "dry_run": dry_run,
        "use_cost": use_cost,
        "elapsed_seconds": round(time.time() - started, 2),
        "cmd": cmd,
        "returncode": result.returncode,
        "stdout": result.stdout,
        "stderr": result.stderr,
        "csv_file": csv_path.name if csv_path.exists() else None,
    }


@app.get("/run/full")
def run_full(request: Request, connected_seller_id: int = 1, limit: int = 0, dry_run: bool = True, use_cost: bool = False):
    user = require_user(request)
    require_connected_seller_access(int(user["id"]), connected_seller_id)
    started = time.time()
    results = {}
    r1 = subprocess.run(build_inventory_cmd(connected_seller_id, limit), capture_output=True, text=True)
    results["inventory"] = r1.stdout or r1.stderr
    r2 = subprocess.run(build_rebate_cmd(connected_seller_id, limit), capture_output=True, text=True)
    results["rebate"] = r2.stdout or r2.stderr
    csv_path = build_csv_path()
    r3 = subprocess.run(build_optimizer_cmd(connected_seller_id, limit, dry_run, use_cost, csv_path), capture_output=True, text=True)
    results["optimizer"] = r3.stdout or r3.stderr
    return {
        "status": "full run executado",
        "connected_seller_id": connected_seller_id,
        "limit": limit,
        "dry_run": dry_run,
        "use_cost": use_cost,
        "elapsed_seconds": round(time.time() - started, 2),
        "csv_file": csv_path.name if csv_path.exists() else None,
        "results": results,
    }


def async_ok_response(message: str, run_id: str, connected_seller_id: int, limit: int, **extra):
    payload = {
        "status": message,
        "run_id": run_id,
        "connected_seller_id": connected_seller_id,
        "limit": limit,
    }
    payload.update(extra)
    return JSONResponse(payload)


@app.get("/run/inventory_async")
def run_inventory_async(request: Request, connected_seller_id: int = 1, limit: int = 0):
    user = require_user(request)
    seller_access = require_connected_seller_access(int(user["id"]), connected_seller_id)
    account_id = int(seller_access["account_id"])
    validate_plan_access(account_id, limit, enforce_mlb_limit=False)
    log_path = build_log_path("inventory_run")
    run_id = insert_job(
        job_type="inventory",
        connected_seller_id=connected_seller_id,
        limit_items=limit,
        dry_run=None,
        use_cost=None,
        payload={"cmd": build_inventory_cmd(connected_seller_id, limit)},
        log_file=log_path.name,
        csv_file=None,
    )
    return async_ok_response("inventory enfileirado", run_id, connected_seller_id, limit, log_file=log_path.name)


@app.get("/run/rebate_async")
def run_rebate_async(request: Request, connected_seller_id: int = 1, limit: int = 0):
    user = require_user(request)
    seller_access = require_connected_seller_access(int(user["id"]), connected_seller_id)
    account_id = int(seller_access["account_id"])
    validate_plan_access(account_id, limit, enforce_mlb_limit=True)
    log_path = build_log_path("rebate_run")
    run_id = insert_job(
        job_type="rebate",
        connected_seller_id=connected_seller_id,
        limit_items=limit,
        dry_run=None,
        use_cost=None,
        payload={"cmd": build_rebate_cmd(connected_seller_id, limit)},
        log_file=log_path.name,
        csv_file=None,
    )
    return async_ok_response("rebate enfileirado", run_id, connected_seller_id, limit, log_file=log_path.name)


@app.get("/run/optimizer_async")
def run_optimizer_async(request: Request, connected_seller_id: int = 1, limit: int = 0, dry_run: bool = True, use_cost: bool = False):
    user = require_user(request)
    seller_access = require_connected_seller_access(int(user["id"]), connected_seller_id)
    account_id = int(seller_access["account_id"])
    validate_plan_access(account_id, limit, enforce_mlb_limit=True)
    log_path = build_log_path("optimizer_run")
    csv_path = build_csv_path()
    run_id = insert_job(
        job_type="optimizer",
        connected_seller_id=connected_seller_id,
        limit_items=limit,
        dry_run=dry_run,
        use_cost=use_cost,
        payload={
            "rebate_cmd": build_rebate_cmd(connected_seller_id, limit),
            "cmd": build_optimizer_cmd(connected_seller_id, limit, dry_run, use_cost, csv_path),
        },
        log_file=log_path.name,
        csv_file=csv_path.name,
    )
    register_daily_execution(account_id)
    return async_ok_response(
        "optimizer enfileirado",
        run_id,
        connected_seller_id,
        limit,
        dry_run=dry_run,
        use_cost=use_cost,
        log_file=log_path.name,
        csv_file=csv_path.name,
    )


@app.get("/run/full_async")
def run_full_async(request: Request, connected_seller_id: int = 1, limit: int = 0, dry_run: bool = True, use_cost: bool = False):
    user = require_user(request)
    seller_access = require_connected_seller_access(int(user["id"]), connected_seller_id)
    account_id = int(seller_access["account_id"])
    validate_plan_access(account_id, limit, enforce_mlb_limit=True)
    log_path = build_log_path("full_pipeline")
    csv_path = build_csv_path()
    run_id = insert_job(
        job_type="full",
        connected_seller_id=connected_seller_id,
        limit_items=limit,
        dry_run=dry_run,
        use_cost=use_cost,
        payload={
            "inventory_cmd": build_inventory_cmd(connected_seller_id, limit),
            "rebate_cmd": build_rebate_cmd(connected_seller_id, limit),
            "optimizer_cmd": build_optimizer_cmd(connected_seller_id, limit, dry_run, use_cost, csv_path),
        },
        log_file=log_path.name,
        csv_file=csv_path.name,
    )
    register_daily_execution(account_id)
    return async_ok_response(
        "pipeline completo enfileirado",
        run_id,
        connected_seller_id,
        limit,
        dry_run=dry_run,
        use_cost=use_cost,
        log_file=log_path.name,
        csv_file=csv_path.name,
    )


@app.get("/run/status")
def run_status(request: Request, run_id: str):
    user = require_user(request)
    require_run_access(int(user["id"]), run_id)
    return get_job(run_id)


@app.get("/run/log")
def run_log(request: Request, run_id: str):
    user = require_user(request)
    require_run_access(int(user["id"]), run_id)
    job = get_job(run_id)

    log_file = job.get("log_file")
    if log_file:
        path = LOG_DIR / log_file
        if path.exists() and path.is_file():
            text = path.read_text(encoding="utf-8")
            return PlainTextResponse(text)

    result_json = job.get("result_json") or {}
    parts = []

    if isinstance(result_json, dict):
        for step_name in ("inventory", "rebate", "optimizer"):
            step_data = result_json.get(step_name)
            if not isinstance(step_data, dict):
                continue
            stdout = (step_data.get("stdout") or "").strip()
            stderr = (step_data.get("stderr") or "").strip()
            returncode = step_data.get("returncode")
            elapsed = step_data.get("elapsed_seconds")
            parts.append(f"===== {step_name.upper()} | returncode={returncode} | elapsed_seconds={elapsed} =====")
            if stdout:
                parts.append("")
                parts.append("[STDOUT]")
                parts.append("")
                parts.append(stdout)
            if stderr:
                parts.append("")
                parts.append("[STDERR]")
                parts.append("")
                parts.append(stderr)
            parts.append("")

    if parts:
        return PlainTextResponse("\n".join(parts).strip())

    return PlainTextResponse(f"Sem log disponível ainda. status={job.get('status')} step={job.get('step')}")

@app.get("/download/csv")
def download_csv(request: Request, filename: str | None = None, run_id: str | None = None):
    user = require_user(request)
    if run_id:
        require_run_access(int(user["id"]), run_id)
    # caminho 1: arquivo local no mesmo serviço
    if filename:
        file_path = CSV_DIR / filename
        if file_path.exists() and file_path.is_file():
            return FileResponse(path=str(file_path), filename=filename, media_type="text/csv")

    # caminho 2: fallback pelo banco (worker e web em serviços diferentes)
    if run_id:
        payload = get_job_csv_payload(run_id)
        if not payload:
            raise HTTPException(status_code=404, detail="run_id não encontrado")

        csv_content = payload.get("csv_content")
        csv_file = payload.get("csv_file") or "resultado.csv"
        mime = payload.get("csv_mime_type") or "text/csv"

        if csv_content:
            headers = {"Content-Disposition": f'attachment; filename="{csv_file}"'}
            return PlainTextResponse(csv_content, media_type=mime, headers=headers)

    raise HTTPException(status_code=404, detail="CSV não encontrado")


@app.get("/download/log")
def download_log(request: Request, filename: str):
    user = require_user(request)
    file_path = LOG_DIR / filename
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="Log não encontrado")
    return FileResponse(path=str(file_path), filename=filename, media_type="text/plain")


@app.get("/jobs/recent")
def recent_jobs(request: Request, limit: int = 20, connected_seller_id: int | None = None):
    user = require_user(request)
    if connected_seller_id is not None:
        require_connected_seller_access(int(user["id"]), connected_seller_id)
    else:
        connected_seller_id = get_default_connected_seller_id_for_user(int(user["id"]))
    return {"jobs": recent_jobs_rows(limit=limit, connected_seller_id=connected_seller_id)}


@app.get("/jobs/active")
def active_job(request: Request, connected_seller_id: int):
    user = require_user(request)
    require_connected_seller_access(int(user["id"]), connected_seller_id)
    active = get_active_job_for_seller(connected_seller_id)
    if not active:
        return {"active": None}
    active["summary"] = build_job_summary(active)
    return {"active": active}


@app.get("/painel", response_class=HTMLResponse)
def painel(request: Request, connected_seller_id: int | None = None, connected: int = 0, account_id: int | None = None):
    user = require_user(request)
    accessible_sellers = get_accessible_connected_sellers(int(user["id"]))
    if not accessible_sellers:
        return RedirectResponse(url="/onboarding")
    if connected_seller_id is None:
        connected_seller_id = int(accessible_sellers[0]["connected_seller_id"])
    else:
        require_connected_seller_access(int(user["id"]), connected_seller_id)
    seller = get_connected_seller_summary(connected_seller_id)
    current_account_id = int(seller.get("account_id") or accessible_sellers[0]["account_id"])
    current_user_role = get_user_role_for_account(int(user["id"]), current_account_id)
    can_manage_access = current_user_role in ("owner", "admin")

    subscription = get_or_create_subscription(current_account_id)
    plan_name = "-"
    plan_status = "inactive"
    plan_days_left = None
    executions_today = 0
    executions_remaining_today = 0
    daily_execution_limit = None
    mlb_limit_label = "-"
    default_limit = 0
    limit_label = "Todos os anúncios"

    inventory_item_count = get_latest_inventory_item_count(connected_seller_id)
    inventory_plan_warning = ""

    if subscription:
        subscription = expire_trial_if_needed(subscription)
        if subscription and subscription.get("status") != "expired":
            plan_name = subscription.get("plan_name") or subscription.get("plan_code") or "-"
            plan_status = subscription.get("status") or "inactive"
            if subscription.get("mlb_limit") is None:
                default_limit = 0
                mlb_limit_label = "Ilimitado"
                limit_label = "Todos os anúncios"
            else:
                default_limit = int(subscription.get("mlb_limit") or 0)
                mlb_limit_label = str(default_limit)
                limit_label = f"{default_limit} MLBs"

                if inventory_item_count is not None and inventory_item_count > default_limit:
                    excedente = inventory_item_count - default_limit
                    inventory_plan_warning = (
                        f"Sua conta tem {inventory_item_count} MLBs ativos mapeados no último inventory, "
                        f"acima do limite do plano atual ({default_limit}). "
                        f"Rebate, Optimizer e Otimização Completa serão limitados ao plano. "
                        f"Excedente estimado: {excedente} MLBs."
                    )

            daily_execution_limit = subscription.get("daily_execution_limit")
            usage_today = get_today_usage(current_account_id)
            executions_today = int(usage_today.get("executions_count") or 0)
            if daily_execution_limit is not None:
                executions_remaining_today = max(int(daily_execution_limit) - executions_today, 0)

            if subscription.get("current_period_end"):
                now_utc = datetime.now(timezone.utc)
                delta = subscription["current_period_end"] - now_utc
                plan_days_left = max(delta.days, 0)

    if seller["connected"]:
        status_html = f"""
        <div class="status-card connected">
            <div class="status-title">✅ Conta conectada</div>
            <div class="status-line"><strong>Conta:</strong> {seller.get('seller_nickname') or '-'}</div>
            <div class="status-line"><strong>ML User ID:</strong> {seller.get('ml_user_id') or '-'}</div>
            <div class="status-line"><strong>Site:</strong> {seller.get('site_id') or '-'}</div>
        </div>
        """
    else:
        status_html = """
        <div class="status-card disconnected">
            <div class="status-title">❌ Não conectado</div>
            <div class="status-line">Conecte sua conta do Mercado Livre para liberar a execução.</div>
        </div>
        """

    
    trial_banner = ""
    expired_banner = ""

    if plan_status == "trialing" and plan_days_left is not None:
        trial_banner = f'''
        <div style="margin-bottom:16px;padding:14px 18px;border-radius:14px;
        background:rgba(47,212,199,.15);border:1px solid rgba(47,212,199,.35);
        color:#d8f4f1;font-weight:700;text-align:center;">
        🎉 Seu teste grátis está ativo • {plan_days_left} dias restantes
        </div>
        '''

    if plan_status == "expired":
        expired_banner = '''
        <div style="margin-bottom:16px;padding:14px;border-radius:12px;
        background:rgba(239,68,68,.2);border:1px solid rgba(239,68,68,.4);
        color:#fecaca;text-align:center;font-weight:700;">
        🚫 Seu teste expirou • Assine um plano para continuar
        </div>
        '''

    connected_banner = '<div class="flash success">Conta conectada com sucesso.</div>' if connected == 1 else ''
    #new_connect_href = f"/ml/oauth/start?account_id={account_id}" if account_id else "#"
    new_connect_href = f"/ml/oauth/start?account_id={current_account_id}"
    reconnect_href = f"/ml/oauth/start?connected_seller_id={connected_seller_id}"

    return f"""
    <html>
    <head>
        <title>Exos Profit</title>
        <meta name="viewport" content="width=device-width, initial-scale=1" />
        <style>
            * {{ box-sizing: border-box; }}
            body {{ margin: 0; font-family: Arial, sans-serif; background: linear-gradient(180deg, #06122b 0%, #091a3f 100%); color: #ffffff; }}
            .container {{ max-width: 1240px; margin: 0 auto; padding: 32px 20px 56px; }}
            .hero {{ text-align: center; margin-bottom: 24px; }}
            .hero h1 {{ font-size: 48px; margin: 0 0 10px; font-weight: 800; }}
            .hero p {{ font-size: 18px; color: #d8e3ff; margin: 0; }}
            .flash {{ max-width: 760px; margin: 0 auto 20px; padding: 14px 18px; border-radius: 14px; text-align: center; font-weight: 700; }}
            .flash.success {{ background: rgba(34, 197, 94, 0.18); border: 1px solid rgba(34, 197, 94, 0.35); color: #d8ffe5; }}
            .grid {{ display: grid; grid-template-columns: 1.05fr 1fr; gap: 20px; align-items: start; }}
            .card {{ background: rgba(255,255,255,0.08); border: 1px solid rgba(255,255,255,0.08); border-radius: 20px; padding: 24px; box-shadow: 0 10px 30px rgba(0,0,0,0.22); }}
            .card h2 {{ margin: 0 0 18px; font-size: 28px; }}
            .status-card {{ border-radius: 16px; padding: 18px; }}
            .status-card.connected {{ background: rgba(34, 197, 94, 0.14); border: 1px solid rgba(34, 197, 94, 0.28); }}
            .status-card.disconnected {{ background: rgba(239, 68, 68, 0.12); border: 1px solid rgba(239, 68, 68, 0.25); }}
            .status-title {{ font-size: 22px; font-weight: 800; margin-bottom: 8px; }}
            .status-line {{ font-size: 15px; color: #e8eeff; margin-top: 6px; }}
            .actions {{ display: flex; flex-direction: column; gap: 12px; margin-top: 16px; }}
            .btn {{ width: 100%; border: none; border-radius: 14px; padding: 14px 16px; font-size: 16px; font-weight: 700; cursor: pointer; transition: transform .05s ease, opacity .2s ease; text-decoration: none; }}
            .btn:hover {{ opacity: .94; }}
            .btn:active {{ transform: scale(0.99); }}
            .btn-connect {{ background: #22c55e; color: #051b0d; }}
            .btn-primary {{ background: #3b82f6; color: #ffffff; }}
            .btn-secondary {{ background: #0f172a; color: #ffffff; border: 1px solid rgba(255,255,255,0.15); }}
            .btn-warn {{ background: #f59e0b; color: #1a1200; }}
            .btn-danger {{ background: #ef4444; color: #fff; }}
            .form-row {{ margin-bottom: 14px; text-align: left; }}
            .form-row label {{ display: block; margin-bottom: 8px; font-weight: 700; color: #dbe6ff; }}
            input[type="number"] {{ width: 220px; padding: 12px 14px; border-radius: 12px; border: none; font-size: 18px; }}
            .check {{ display: flex; align-items: center; gap: 10px; font-size: 18px; margin: 10px 0; }}
            .output-wrap {{ margin-top: 24px; }}
            .output-head {{ display: flex; justify-content: space-between; align-items: center; gap: 12px; flex-wrap: wrap; margin-bottom: 10px; }}
            .download-area {{ display: none; gap: 10px; flex-wrap: wrap; }}
            .download-area.show {{ display: flex; }}
            pre {{ white-space: pre-wrap; word-break: break-word; text-align: left; background: #020817; color: #dbeafe; border: 1px solid rgba(255,255,255,0.08); padding: 18px; border-radius: 16px; min-height: 260px; max-height: 640px; overflow: auto; font-size: 13px; line-height: 1.45; }}
            .muted {{ color: #9fb0d9; font-size: 14px; margin-top: 8px; }}
            .small-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }}
            .job-list {{ margin-top: 16px; font-size: 13px; color: #d8e3ff; }}
            .job-item {{ padding: 12px 14px; border-radius: 12px; background: rgba(2,8,23,0.55); margin-bottom: 10px; }}
            .job-head {{ display:flex; justify-content: space-between; gap:10px; flex-wrap: wrap; }}
            .badge {{ display:inline-block; padding: 4px 8px; border-radius: 999px; font-size: 12px; font-weight: 700; }}
            .badge.queued {{ background: rgba(59,130,246,.18); color: #bfdbfe; border: 1px solid rgba(59,130,246,.35); }}
            .badge.running {{ background: rgba(245,158,11,.18); color: #fde68a; border: 1px solid rgba(245,158,11,.35); }}
            .badge.finished {{ background: rgba(34,197,94,.18); color: #bbf7d0; border: 1px solid rgba(34,197,94,.35); }}
            .badge.error {{ background: rgba(239,68,68,.18); color: #fecaca; border: 1px solid rgba(239,68,68,.35); }}
            .summary-card {{ background: rgba(2,8,23,0.58); border: 1px solid rgba(255,255,255,0.06); border-radius: 16px; padding: 16px; margin-top: 16px; }}
            .summary-title {{ font-size: 18px; font-weight: 800; margin-bottom: 8px; }}
            .summary-headline {{ font-size: 16px; color: #e8eeff; }}
            .metrics {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin-top: 14px; }}
            .metric {{ background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.06); border-radius: 14px; padding: 12px; }}
            .metric-label {{ font-size: 12px; color: #9fb0d9; margin-bottom: 4px; }}
            .metric-value {{ font-size: 22px; font-weight: 800; }}
            .warn-box {{ display:none; margin-top: 10px; padding: 12px 14px; border-radius: 12px; background: rgba(245,158,11,.14); border:1px solid rgba(245,158,11,.28); color:#fde68a; }}
            .plan-card {{ background: rgba(2,8,23,0.58); border: 1px solid rgba(255,255,255,0.08); border-radius: 16px; padding: 16px; margin: 14px 0 0; }}
            .plan-label {{ font-size: 12px; color: #9fb0d9; margin-bottom: 6px; }}
            .plan-name {{ font-size: 22px; font-weight: 800; margin-bottom: 8px; }}
            .plan-meta {{ font-size: 13px; color: #d8e3ff; margin-top: 4px; }}
            .status-pill {{ display:inline-block; padding: 4px 10px; border-radius: 999px; font-size: 12px; font-weight: 700; }}
            .status-pill.trialing {{ background: rgba(59,130,246,.18); color: #bfdbfe; border: 1px solid rgba(59,130,246,.35); }}
            .status-pill.active {{ background: rgba(34,197,94,.18); color: #bbf7d0; border: 1px solid rgba(34,197,94,.35); }}
            .status-pill.past_due {{ background: rgba(245,158,11,.18); color: #fde68a; border: 1px solid rgba(245,158,11,.35); }}
            .status-pill.paused {{ background: rgba(148,163,184,.18); color: #e2e8f0; border: 1px solid rgba(148,163,184,.35); }}
            .status-pill.inactive, .status-pill.expired, .status-pill.canceled {{ background: rgba(239,68,68,.18); color: #fecaca; border: 1px solid rgba(239,68,68,.35); }}
            .scope-box {{ margin-top: 6px; padding: 12px 14px; border-radius: 12px; background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.06); color:#ffffff; font-weight:700; }}
            .scope-help {{ font-size: 12px; color: #9fb0d9; margin-top: 6px; }}
            #customLimitRow {{ display: none; }}
            a.button-link {{ text-decoration: none; display: block; }}
            .invite-row {{ display:flex; gap:10px; flex-wrap:wrap; align-items:end; margin-top:12px; }}
            #inviteRole {{
                width: 160px;
                height: 40px;
                padding: 0 12px;
                border-radius: 10px;
                border: none;
                font-size: 14px;
                line-height: 40px;
                background: #f8fafc;
                color: #111827;
            }}

            #inviteEmail {{
                width: 360px;
                max-width: 100%;
                height: 40px;
                padding: 0 12px;
                border-radius: 10px;
                border: none;
                font-size: 14px;
                line-height: 40px;
            }}

            .invite-row button {{
                height: 40px;
                padding: 0 18px;
                font-size: 14px;
            }}


            .invite-row > div {{ display:flex; flex-direction:column; }}
            .invite-row label {{ display:block; margin-bottom:6px; font-size:13px; font-weight:700; color:#dbe6ff; }}
            #inviteEmail {{ width: 340px; max-width: 100%; height: 40px; padding: 0 12px; border-radius: 10px; border: none; font-size: 14px; line-height: 40px; }}
            #inviteRole {{ width: 150px; height: 40px; padding: 0 12px; border-radius: 10px; border: none; font-size: 14px; line-height: 40px; background: #f8fafc; color: #111827; }}
            .invite-list {{ margin-top: 14px; }}
            .invite-item {{ padding: 10px 12px; border-radius: 12px; background: rgba(2,8,23,0.55); margin-bottom: 8px; display:flex; justify-content:space-between; gap:12px; flex-wrap:wrap; align-items:center; }}
            .invite-meta {{ color:#9fb0d9; font-size:12px; margin-top:4px; }}
            .topbar {{ display:flex; justify-content:space-between; align-items:center; gap:12px; flex-wrap:wrap; margin-bottom:18px; }}
            .card.compact {{ padding: 18px; }}
            .user-pill {{ padding:8px 12px; border-radius:999px; background: rgba(255,255,255,0.08); color:#dbeafe; font-size:13px; }}
            @media (max-width: 980px) {{ .grid {{ grid-template-columns: 1fr; }} .small-grid {{ grid-template-columns: 1fr; }} .metrics {{ grid-template-columns: 1fr 1fr; }} .hero h1 {{ font-size: 40px; }} }}
            .card.compact {{padding: 16px;min-height: unset;}}
</style>
    </head>
    <body>
        <div class="container">
            <div class="topbar">
                <div class="user-pill">Logado como: {user.get('email')}</div>
                <button class="btn btn-secondary" style="width:auto;" onclick="logout()">Sair</button>
            </div>
            <div class="hero">
                <h1>🚀 Exos Tools</h1>
                <p>Automação de campanhas do Mercado Livre com execução segura e histórico de jobs.</p>
                <div class="muted" style="margin-top:10px;">Usuário: {user.get('full_name') or user.get('email')}</div>
                <div style="margin-top:10px;"><button class="btn btn-secondary" style="width:auto; padding:10px 14px; font-size:14px;" onclick="logout()">Sair</button></div>
            </div>
            {connected_banner}{trial_banner}{expired_banner}
            <div class="grid">
                <div class="card">
                    <h2>Conta Mercado Livre</h2>
                    {status_html}
                    <div class="plan-card">
                        <div class="plan-label">Plano atual</div>
                        <div class="plan-name">{plan_name}</div>
                        <div class="plan-meta"><span class="status-pill {plan_status}">{plan_status.upper()}</span></div>
                        <div class="plan-meta">Limite por execução: {mlb_limit_label} MLBs</div>
                        <div class="plan-meta">Execuções hoje: {executions_today}{f' / {daily_execution_limit}' if daily_execution_limit is not None else ''}</div>
                        <div class="plan-meta">Execuções restantes hoje: {executions_remaining_today}</div>
                        {f'<div class="plan-meta" style="color:#86efac;">{plan_days_left} dias restantes no período atual</div>' if plan_days_left is not None and plan_status == 'trialing' else ''}
                        {f'<div class="plan-meta" style="margin-top:10px; padding:10px 12px; border-radius:12px; background:rgba(245,158,11,.14); border:1px solid rgba(245,158,11,.28); color:#fde68a;">{inventory_plan_warning}</div>' if inventory_plan_warning else ''}
                    </div>
                    <div class="actions">
                        <a class="button-link" href="{new_connect_href}" onclick="return validarNovaConta();"><button class="btn btn-connect" type="button">Conectar nova conta Mercado Livre</button></a>
                        <a class="button-link" href="{reconnect_href}"><button class="btn btn-secondary" type="button">Reconectar seller atual</button></a>
                    </div>
                    <div class="job-list">
                        <div><strong>Histórico recente</strong></div>
                        <div id="recentJobs">Carregando...</div>
                    </div>
                </div>
                <div style="display:flex; flex-direction:column; gap:20px;">
                <div class="card">
                    <h2>Executar otimização</h2>
                    <div class="form-row">
                        <label>Conta ativa</label>
                        <div class="status-line"><strong>{seller.get('seller_nickname') or 'Conta conectada'}</strong></div>
                        <div class="muted">Esta otimização será executada para a conta conectada acima.</div>
                        <input type="hidden" id="connectedSellerId" value="{connected_seller_id}" />
                    </div>
                    <div class="form-row">
                        <label>Escopo da execução</label>
                        <div class="scope-box">{limit_label}</div>
                        <div class="scope-help">Rebate, Optimizer e Otimização Completa usam o escopo do plano. Inventory pode mapear toda a conta para sinalizar excedentes.</div>
                        <input type="hidden" id="limit" value="{default_limit}" />
                    </div>
                    <div class="check"><input type="checkbox" id="dryrun" checked /><label for="dryrun">Simular antes de aplicar</label></div>
                    <div class="muted" style="margin-top:5px; line-height:1.5;">Se Simular antes de aplicar estiver marcado, nada será alterado no Mercado Livre.</div>
                    <div class="check"><input type="checkbox" id="usecost" /><label for="usecost">Usar custo do produto</label></div>
                    <div class="warn-box" id="activeJobWarn" style="display:none;"></div>
                    <div class="actions">
                        <div class="small-grid">
                            <button class="btn btn-secondary" onclick="rodarInventoryAsync()">Atualizar anúncios e fretes</button>
                            <button class="btn btn-primary" onclick="rodarOptimizerAsync()">Aplicar melhor campanha</button>
                        </div>
                        <div class="muted" style="margin-top:5px; line-height:1.5;">
                            Aplicar melhor campanha atualiza os rebates/campanhas disponíveis e depois roda o optimizer.
                        </div>
                        <div class="small-grid" style="margin-top:12px;">
                            <button class="btn btn-connect" onclick="rodarFullAsync()">Rodar atualização completa</button>
                        </div>
                        <div class="muted" style="margin-top:5px; line-height:1.5;">
                            Rodar atualização completa executa todo o processo: atualiza anúncios e fretes, atualiza rebates/campanhas disponíveis e depois ativa melhor campanha.
                        </div>
                    </div>
                    <!--<div class="muted" id="jobInfo"></div>-->
                    <div class="summary-card">
                        <div class="summary-title">Resumo geral</div>
                        <div class="summary-headline" id="summaryHeadline">Nenhum job executado ainda.</div>
                        <div class="muted" id="summaryDetails"></div>
                        <div class="metrics">
                            <div class="metric"><div class="metric-label">Itens analisados</div><div class="metric-value" id="mProcessed">-</div></div>
                            <div class="metric"><div class="metric-label">Trocas elegíveis</div><div class="metric-value" id="mSwitched">-</div></div>
                            <div class="metric"><div class="metric-label">Sem ação</div><div class="metric-value" id="mNoAction">-</div></div>
                            <div class="metric"><div class="metric-label">Erros/Falhas</div><div class="metric-value" id="mErrors">-</div></div>
                        </div>
                    </div>
                </div>
                <div class="card">
                    <h2>Valor Mínimo a receber por SKU</h2>
                    <div class="muted">Baixe o template, preencha <strong>sku</strong> e <strong>vlr_min_receber</strong> e envie o arquivo.</div>
                    <input type="hidden" id="currentAccountId" value="{current_account_id}" />
                    <div class="actions" style="margin-top:14px;">
                        <a class="button-link" href="/template/sku-min-receber.csv" target="_blank"><button class="btn btn-secondary" type="button">Baixar template CSV</button></a>
                    </div>
                    <div class="invite-row" style="margin-top:14px;">
                        <div>
                            <label for="minReceiveFile">Arquivo CSV ou XLSX</label>
                            <input type="file" id="minReceiveFile" accept=".csv,.xlsx" />
                        </div>
                        <div>
                            <button class="btn btn-primary" style="width:auto; height:40px; padding:0 18px; font-size:14px;" onclick="uploadMinReceive()">Enviar arquivo</button>
                        </div>
                    </div>
                    <div class="muted" id="minReceiveInfo" style="margin-top:12px;">Nenhum arquivo enviado nesta sessão.</div>
                    <div class="invite-list" id="minReceiveList">Carregando tabela...</div>
                </div>
                </div>
            </div>
            <div class="card" style="margin-top:20px;">
                <h2>Acessos da conta</h2>
                <input type="hidden" id="currentAccountId" value="{current_account_id}" />
                <div class="muted">Conta atual: {seller.get('seller_nickname') or '-'} | Perfil: {current_user_role or '-'}</div>
                {'' if can_manage_access else '<div class="muted" style="color:#fde68a;">Você não tem permissão para gerenciar acessos desta conta.</div>'}
                <div id="inviteManager" {'style="display:none;"' if not can_manage_access else ''}>
                    <div class="invite-row">
                        <div>
                            <label for="inviteEmail">E-mail para liberar acesso</label>
                            <input type="email" id="inviteEmail" placeholder="cliente@gmail.com" />
                        </div>
                        <div>
                            <label for="inviteRole">Perfil</label>
                            <select id="inviteRole">
                                <option value="owner">Dono</option>
                                <option value="admin">Administrador</option>
                                <option value="viewer">Somente leitura</option>
                            </select>
                        </div>
                        <div>
                            <button class="btn btn-primary" style="width:auto; height:40px; padding:0 18px; font-size:14px;" onclick="criarConvite()">Liberar acesso</button>
                        </div>
                    </div>
                </div>
                <div class="invite-list" id="inviteList">Carregando acessos...</div>
            </div>
            <div class="output-wrap">
                <div class="output-head">
                    <h2>Resultado / Log</h2>
                    <div id="downloadArea" class="download-area">
                        <a id="downloadCsvLink" href="#" target="_blank"><button class="btn btn-secondary" type="button">Baixar CSV</button></a>
                                            </div>
                </div>
                <pre id="output">Nenhuma execução ainda.</pre>
            </div>
        </div>
        <script>
            let pollingTimer = null;

            function validarNovaConta() {{
                const hasAccount = {str(bool(account_id)).lower()};
                if (!hasAccount) {{ alert("Abra o painel com ?account_id=SEU_ID antes de conectar uma nova conta."); return false; }}
                return true;
            }}

            function toggleLimitInput() {{}}

            function getLimitValue() {{
                return document.getElementById("limit").value || 0;
            }}

            function getParams() {{
                return {{ connectedSellerId: document.getElementById("connectedSellerId").value, limit: getLimitValue(), dryRun: document.getElementById("dryrun").checked, useCost: document.getElementById("usecost").checked }};
            }}

            function setOutput(text) {{ document.getElementById("output").innerText = text; }}
            function setJobInfo(text) {{ document.getElementById("jobInfo").innerText = text || ""; }}

            function setDownloads(csvFile, runId, hasCsv, statusValue) {{
                const area = document.getElementById("downloadArea");
                const csvLink = document.getElementById("downloadCsvLink");
                let show = false;

                const finished = statusValue === "finished";
                if (finished && hasCsv && runId) {{
                    csvLink.href = `/download/csv?run_id=${{encodeURIComponent(runId)}}`;
                    csvLink.style.display = "inline-block";
                    show = true;
                }} else {{
                    csvLink.style.display = "none";
                    csvLink.href = "#";
                }}

                area.classList.toggle("show", show);
            }}

            function setSummary(summary) {{
                document.getElementById("summaryHeadline").innerText = summary?.headline || "Sem resumo disponível.";
                document.getElementById("summaryDetails").innerText = (summary?.details || []).join(" | ") || "";
                const m = summary?.metrics || {{}};
                document.getElementById("mProcessed").innerText = m.processed ?? m.item_ids_found ?? "-";
                document.getElementById("mSwitched").innerText = m.switched ?? "-";
                document.getElementById("mNoAction").innerText = m.no_action ?? "-";
                document.getElementById("mErrors").innerText = m.errors ?? m.failed_items ?? m.rebate_failed_items ?? "-";
            }}

            function setActiveWarn(active) {{
                const box = document.getElementById("activeJobWarn");
                if (!active) {{
                    box.style.display = "none";
                    box.innerText = "";
                    return;
                }}
                box.style.display = "block";
                //box.innerText = `Já existe job em andamento: ${{active.job_type}} | status=${{active.status}} | etapa=${{active.step || '-'}} | run_id=${{active.run_id}}`;
                box.innerText = `Já existe job em andamento, aguarde sua finalização`;
            }}

            async function fetchJson(url, options = undefined) {{
                const res = await fetch(url, options);
                const text = await res.text();
                try {{
                    const data = JSON.parse(text);
                    if (!res.ok) {{
                        const detail = typeof data.detail === 'string' ? data.detail : JSON.stringify(data.detail);
                        throw new Error(`HTTP ${{res.status}}\n\n${{detail}}`);
                    }}
                    return data;
                }} catch (e) {{
                    if (e instanceof Error && e.message.startsWith('HTTP ')) throw e;
                    throw new Error(`HTTP ${{res.status}}\n\n${{text}}`);
                }}
            }}

            async function fetchText(url) {{ const res = await fetch(url); return await res.text(); }}
            function stopPolling() {{ if (pollingTimer) {{ clearInterval(pollingTimer); pollingTimer = null; }} }}

            async function logout() {{
                await fetch('/auth/logout', {{ method: 'POST' }});
                window.location.href = '/login';
            }}

            async function refreshInvites() {{
                const accountId = document.getElementById('currentAccountId')?.value;
                const root = document.getElementById('inviteList');
                if (!accountId || !root) return;
                try {{
                    const data = await fetchJson(`/account/invites?account_id=${{encodeURIComponent(accountId)}}`);
                    const invites = data.invites || [];
                    if (!invites.length) {{
                        root.innerHTML = '<div class="muted">Nenhum acesso liberado ainda.</div>';
                        return;
                    }}
                    root.innerHTML = invites.map(i => `
                        <div class="invite-item">
                            <div>
                                <div><strong>${{i.email}}</strong> — ${{(i.role || '-').toUpperCase()}}</div>
                                <div class="invite-meta">status=${{i.status}} | criado em ${{fmtDate(i.created_at)}}${{i.accepted_at ? ` | aceito em ${{fmtDate(i.accepted_at)}}` : ''}}</div>
                            </div>
                            ${{i.status !== 'revoked' ? `<button class="btn btn-secondary" style="width:auto; padding:8px 12px; font-size:13px;" onclick="revogarConvite(${{i.id}})">Revogar</button>` : ''}}
                        </div>
                    `).join('');
                }} catch (e) {{
                    root.innerHTML = `<div class="muted">${{String(e)}}</div>`;
                }}
            }}

            async function criarConvite() {{
                const accountId = document.getElementById('currentAccountId').value;
                const email = (document.getElementById('inviteEmail').value || '').trim();
                const role = document.getElementById('inviteRole').value;
                if (!email) {{
                    alert('Informe o e-mail do usuário.');
                    return;
                }}
                try {{
                    const url = `/account/invites/create?account_id=${{encodeURIComponent(accountId)}}&email=${{encodeURIComponent(email)}}&role=${{encodeURIComponent(role)}}`;
                    await fetchJson(url, {{ method: 'POST' }});
                    document.getElementById('inviteEmail').value = '';
                    await refreshInvites();
                    alert('Acesso liberado com sucesso.');
                }} catch (e) {{
                    alert(String(e));
                }}
            }}

            async function revogarConvite(inviteId) {{
                const accountId = document.getElementById('currentAccountId').value;
                if (!confirm('Revogar este acesso?')) return;
                try {{
                    const url = `/account/invites/revoke?account_id=${{encodeURIComponent(accountId)}}&invite_id=${{encodeURIComponent(inviteId)}}`;
                    await fetchJson(url, {{ method: 'POST' }});
                    await refreshInvites();
                }} catch (e) {{
                    alert(String(e));
                }}
            }}

            function badgeHtml(status) {{
                const s = (status || "").toLowerCase();
                const label = (status || "-").toUpperCase();
                return `<span class="badge ${{s}}">${{label}}</span>`;
            }}

            function fmtDate(value) {{
                if (!value) return "-";
                try {{
                    return new Date(value).toLocaleString("pt-BR");
                }} catch (e) {{
                    return value;
                }}
            }}

            async function refreshRecentJobs() {{
                try {{
                    const sellerId = document.getElementById("connectedSellerId").value;
                    const data = await fetchJson(`/jobs/recent?connected_seller_id=${{sellerId}}&limit=8`);
                    const active = await fetchJson(`/jobs/active?connected_seller_id=${{sellerId}}`);
                    setActiveWarn(active.active);
                    const root = document.getElementById("recentJobs");
                    const jobs = data.jobs || [];
                    if (!jobs.length) {{ root.innerHTML = "Nenhum job recente."; return; }}
                    root.innerHTML = jobs.map(j => `
                        <div class="job-item">
                            <div class="job-head">
                                <div><strong>${{j.job_type}}</strong> ${{badgeHtml(j.status)}}</div>
                                <div>${{j.step || '-'}}</div>
                            </div>
                            <div style="margin-top:6px">${{j.summary?.headline || ''}}</div>
                            <!--<div style="margin-top:6px; color:#9fb0d9">run_id=${{j.run_id}}</div>-->
                            <!--<div style="margin-top:4px; color:#9fb0d9">criado_em=${{fmtDate(j.created_at)}}</div>-->
                            <div style="margin-top:8px; display:flex; gap:8px; flex-wrap:wrap;">
                                <button class="btn btn-secondary" style="width:auto; padding:8px 12px; font-size:13px;" onclick="verJob('${{j.run_id}}')">Ver job</button>
                                <a target="_blank" href="/run/status?run_id=${{j.run_id}}"><button class="btn btn-secondary" style="width:auto; padding:8px 12px; font-size:13px;" type="button">Status</button></a>
                                <a target="_blank" href="/run/log?run_id=${{j.run_id}}"><button class="btn btn-secondary" style="width:auto; padding:8px 12px; font-size:13px;" type="button">Log</button></a>
                                ${{j.has_csv ? `<a target="_blank" href="/download/csv?run_id=${{encodeURIComponent(j.run_id)}}"><button class="btn btn-secondary" style="width:auto; padding:8px 12px; font-size:13px;" type="button">CSV</button></a>` : ''}}
                            </div>
                        </div>
                    `).join("");
                }} catch (e) {{ document.getElementById("recentJobs").innerText = String(e); }}
            }}

            async function pollRun(runId) {{
                try {{
                    const status = await fetchJson(`/run/status?run_id=${{encodeURIComponent(runId)}}`);
                    const logText = await fetchText(`/run/log?run_id=${{encodeURIComponent(runId)}}`);
                    setOutput(logText || JSON.stringify(status, null, 2));
                    setJobInfo(`run_id=${{status.run_id}} | tipo=${{status.job_type}} | status=${{status.status}} | etapa=${{status.step || '-'}} | iniciado=${{status.started_at || '-'}} | fim=${{status.finished_at || '-'}}`);
                    setDownloads(status.csv_file, status.run_id, status.has_csv, status.status);
                    setSummary(status.summary);
                    await refreshRecentJobs();
                    if (status.status === "finished" || status.status === "error") stopPolling();
                }} catch (err) {{ setOutput(String(err)); stopPolling(); }}
            }}

            function startPolling(runId) {{ stopPolling(); pollRun(runId); pollingTimer = setInterval(() => pollRun(runId), 3000); }}
            async function verJob(runId) {{ startPolling(runId); }}

            async function runAsync(url, message) {{
                stopPolling();
                setOutput(message);
                try {{
                    const data = await fetchJson(url);
                    setOutput(JSON.stringify(data, null, 2));
                    setJobInfo(`run_id=${{data.run_id}} | status=${{data.status}}`);
                    setDownloads(data.csv_file || null, data.run_id || null, false, "queued");
                    setSummary({{ headline: 'Job enfileirado, aguardando processamento.', metrics: {{}} }});
                    await refreshRecentJobs();
                    if (data.run_id) startPolling(data.run_id);
                }} catch (err) {{
                    setOutput(String(err));
                    await refreshRecentJobs();
                }}
            }}

            async function rodarInventoryAsync() {{ const p = getParams(); await runAsync(`/run/inventory_async?connected_seller_id=${{p.connectedSellerId}}&limit=${{p.limit}}`, "Enfileirando inventory..."); }}
            async function rodarRebateAsync() {{ const p = getParams(); await runAsync(`/run/rebate_async?connected_seller_id=${{p.connectedSellerId}}&limit=${{p.limit}}`, "Enfileirando rebate..."); }}
            async function rodarOptimizerAsync() {{ const p = getParams(); await runAsync(`/run/optimizer_async?connected_seller_id=${{p.connectedSellerId}}&limit=${{p.limit}}&dry_run=${{p.dryRun}}&use_cost=${{p.useCost}}`, "Enfileirando optimizer..."); }}
            async function rodarFullAsync() {{ const p = getParams(); await runAsync(`/run/full_async?connected_seller_id=${{p.connectedSellerId}}&limit=${{p.limit}}&dry_run=${{p.dryRun}}&use_cost=${{p.useCost}}`, "Enfileirando pipeline completa..."); }}

            async function logout() {{
                await fetch("/auth/logout", {{ method: "POST" }});
                window.location.href = "/login";
            }}

            toggleLimitInput();
            refreshRecentJobs();
            refreshInvites();
            refreshMinReceive();
        
function renderActiveJobWarn(data) {{
  const el = document.getElementById('activeJobWarn');

  if (!el) return; // 👈 ESSENCIAL

  let msg = "⚠️ Já existe uma atualização em andamento. Aguarde a conclusão antes de iniciar outra.";

  try {{
    const st = data?.status || data?.detail?.status;

    if (st === 'queued') {{
      msg = "⏳ Sua solicitação está na fila. Em breve será iniciada.";
    }} else if (st === 'running') {{
      msg = "🔄 Estamos processando sua otimização agora.";
    }}
  }} catch(e){{}}

  el.innerText = msg;
  el.style.display = 'block';
}}

</script>
    </body>
    </html>
    """